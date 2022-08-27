import os
import openpyxl

from datetime import datetime 

from instagram_crawling.meta_data import BASE_URL, EXCEL_DIR

def add_value_in_excel_cell(sheet, num, val, key=None):
	''' 해당 셀에 값 추가하는 함수 '''

	if key is not None:
		sheet.cell(row=num, column=1).value = key
	sheet.cell(row=num, column=2).value = val


def import_excel(file_path, import_type):
	''' 결과값을 액셀에서 불러오는 함수 '''

	try:
		# 엑셀 열기
		wb = openpyxl.load_workbook(file_path)

		# sheet 선택
		sheet = wb.get_sheet_by_name(import_type)

		# 엑셀 마지막 row 값 가져오기
		max_row = sheet.max_row

		# 게시글 url
		if import_type == 'post_url':
			result_list = list()
			extract_count = sheet.cell(row=1, column=2).value

			# 두번째 row부터 마지막 row까지의 값 가져오기 
			for i in range(1, max_row):
				result_list.append(sheet.cell(row=i+1, column=2).value)
			
			return {'results': result_list, 'extract_count': extract_count}

		# 해쉬태그
		elif import_type == 'tag_name':
			result_dict = dict()

			# 두번째 row부터 마지막 row까지의 값 가져오기 
			for i in range(1, max_row):
				result_dict[sheet.cell(row=i+1, column=1).value] = sheet.cell(row=i+1, column=2).value

			return {'results': result_dict, 'tag_count': len(result_dict)}

	except (OSError, KeyError) as e:
		return {}

def export_excel(hash_tag, export_type, results, file_path=None):
	''' 결과값을 액셀로 저장하는 함수 '''

	wb = None

	date = datetime.today().strftime('%Y%m%d')
	file_name = f'{hash_tag}_{date}.xlsx'
	file_path = os.path.join(EXCEL_DIR, file_name) if file_path is None else file_path

	# 게시글 url 저장할때
	if export_type == 'post_url':
		# 엑셀 열기
		wb = openpyxl.Workbook()

		# 새시트 활성화
		sheet = wb.active

		# 시트명 입력
		sheet.title = export_type

		# 추출 횟수 저장
		sheet.cell(row=1, column=1).value = '해쉬태그 추출 횟수: '
		sheet.cell(row=1, column=2).value = 0

		# 결과를 엑셀 두번째 row부터 하나씩 저장 (url 주소)
		for i, url in enumerate(results, start=2):
			add_value_in_excel_cell(sheet, i, f'{BASE_URL}{url}')

	# 해쉬태그 결과 저장할때
	elif export_type == 'tag_name':
		# 엑셀 열기
		wb = openpyxl.load_workbook(file_path)

		# 새시트 생성
		sheet = wb.create_sheet()

		# 시트명 입력
		sheet.title = export_type

		# 추출 결과 저장
		sheet.cell(row=1, column=1).value = f'{hash_tag} 추출 결과: '
		sheet.cell(row=1, column=2).value = len(results)

		# 결과를 엑셀 두번째 row부터 하나씩 저장 (해쉬태그와 갯수 정보)
		for i, tag_name in enumerate(list(results.keys()), start=2):
			add_value_in_excel_cell(sheet, i, results[tag_name], tag_name)

	# 저장
	wb.save(file_path)

	return file_path

def update_excel(file_path, update_type, count, results=None):
	''' 결과값을 엑셀에 업데이트하는 함수 '''

	try:
		# 엑셀 열기
		wb = openpyxl.load_workbook(file_path)

		# sheet 선택
		sheet = wb.get_sheet_by_name(update_type)

		# 게시글 url 저장할때
		if update_type == 'post_url':
			# 추출 횟수
			sheet.cell(row=1, column=2).value = int(count)

			# url 주소를 2번째 row부터 업데이트
			if results:
				for i, url in enumerate(results, start=2):
					add_value_in_excel_cell(sheet, i, f'{BASE_URL}{url}')

		# 해쉬태그 결과 저장할때
		elif update_type == 'tag_name':
			sheet.cell(row=1, column=2).value = int(count)

			# 변경된 값 엑셀에 업데이트 (해쉬태그와 갯수 정보)
			for i, tag_name in enumerate(list(results.keys()), start=2):
				add_value_in_excel_cell(sheet, i, results[tag_name], tag_name)

		# 저장
		wb.save(file_path)

	except OSError as e:
		pass